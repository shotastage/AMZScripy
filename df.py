import openpyxl as px


EXCEL_SHEET_COL_IDX = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O",
    "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"
]

class DataFrame():

    def __init__(self):
        self._names = []
        self._val = []

    def add_col(self, name, value):
        self._names.append(name)
        self._val.append(value)


    def save_csv(self, fname):

        with open(fname, mode='w') as f:
            # Write col names
            for i, val in enumerate(self._names):
                if i == len(self._names) - 1:
                    f.write(val + ",,,\n")
                else:
                    f.write(val + ",")

            for i in range(len(self._val[0])):
                for j, val in enumerate(self._val):

                    if j == len(self._names) - 1:
                        f.write(val[i]+",,,\n")
                    else:
                        f.write(val[i]+",")

    def save_excel(self, fname):
        wb = px.Workbook()
        ws = wb.active
        ws.title = "AMZ Result"

        for i, val in enumerate(self._names):
            ws[EXCEL_SHEET_COL_IDX[i]+"1"] = val

        for i in range(len(self._val[0])):
            for j, val in enumerate(self._val):
                ws[EXCEL_SHEET_COL_IDX[j]+str(i + 2)] = val[i]

        wb.save(fname)
